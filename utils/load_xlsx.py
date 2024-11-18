from openpyxl import load_workbook


class FileSearchXlsx:
    def __init__(self):
        """
        Carrega o arquivo xlsx com os códigos de produtos e municípios.

        Attributes
        ----------
        path_xlsx : str
            Caminho do arquivo xlsx.
        workbook : openpyxl.workbook
            Instância do workbook.
        sheet : openpyxl.worksheet
            Instância da sheet ativa.
        """
        self.path_xlsx = './src/data/codigos.xlsx'
        self.workbook = load_workbook(self.path_xlsx)
        self.sheet = self.workbook.active

    def get_column(self, column_name):
        header = [cell.value for cell in self.sheet[1]]

        if column_name not in header:
            raise KeyError(
                f"A coluna '{column_name}' não existe no arquivo Excel.")

        col_index = header.index(column_name) + 1

        column_values = [
            self.sheet.cell(row=row_idx, column=col_index).value
            for row_idx in range(2, self.sheet.max_row + 1)
        ]
        return [val for val in column_values if val is not None]
